"""VAT reporting services."""

import csv
from decimal import Decimal
from datetime import datetime, date
from io import StringIO, BytesIO

from django.conf import settings
from django.core.files.base import ContentFile
from django.utils import timezone

from apps.reporting.models import VATReport, VATTransactionLog, TaxExemption
from apps.orders.models import Order
from apps.stores.models import Store


class VATReportService:
    """Generate monthly VAT reports for tax authority."""

    VAT_RATE = Decimal("0.15")  # 15% in Saudi Arabia

    @staticmethod
    def generate_monthly_report(store: Store, year: int, month: int) -> VATReport:
        """
        Generate VAT report for given month.

        Returns:
            VATReport instance
        """
        # Calculate period
        if month == 12:
            period_start = date(year, month, 1)
            period_end = date(year + 1, 1, 1) - timezone.timedelta(days=1)
        else:
            period_start = date(year, month, 1)
            period_end = date(year, month + 1, 1) - timezone.timedelta(days=1)

        # Check if report already exists
        report_number = f"VAT-{year}-{month:02d}"
        report, created = VATReport.objects.update_or_create(
            store=store,
            period_start=period_start,
            period_end=period_end,
            defaults={
                "report_number": report_number,
                "status": VATReport.REPORT_STATUS_DRAFT,
            },
        )

        # Calculate figures
        VATReportService._calculate_figures(report)

        return report

    @staticmethod
    def _calculate_figures(report: VATReport):
        """Calculate all VAT figures for the report."""
        # Get all orders in period
        orders = Order.objects.filter(
            store=report.store,
            created_at__date__gte=report.period_start,
            created_at__date__lte=report.period_end,
        )

        # Calculate totals
        total_sales = Decimal(0)
        total_vat = Decimal(0)
        total_refunds = Decimal(0)
        refund_vat = Decimal(0)

        for order in orders:
            if order.status != "cancelled":
                total_sales += order.total_amount or Decimal(0)
                total_vat += order.tax_amount or Decimal(0)

                # Log transaction
                VATTransactionLog.objects.get_or_create(
                    report=report,
                    invoice_number=f"ORD-{order.id}",
                    defaults={
                        "transaction_type": VATTransactionLog.TRANSACTION_TYPE_SALE,
                        "transaction_date": order.created_at.date(),
                        "amount_ex_vat": (order.total_amount or Decimal(0))
                        - (order.tax_amount or Decimal(0)),
                        "vat_amount": order.tax_amount or Decimal(0),
                        "amount_inc_vat": order.total_amount or Decimal(0),
                        "payment_method": order.payment_method or "online",
                    },
                )

        # Get refunds
        refunded_orders = Order.objects.filter(
            store=report.store,
            status="refunded",
            refunded_at__date__gte=report.period_start,
            refunded_at__date__lte=report.period_end,
        )

        for order in refunded_orders:
            total_refunds += order.total_amount or Decimal(0)
            refund_vat += order.tax_amount or Decimal(0)

            # Log refund
            VATTransactionLog.objects.get_or_create(
                report=report,
                invoice_number=f"REF-{order.id}",
                defaults={
                    "transaction_type": VATTransactionLog.TRANSACTION_TYPE_REFUND,
                    "transaction_date": order.refunded_at.date(),
                    "amount_ex_vat": (order.total_amount or Decimal(0))
                    - (order.tax_amount or Decimal(0)),
                    "vat_amount": order.tax_amount or Decimal(0),
                    "amount_inc_vat": order.total_amount or Decimal(0),
                },
            )

        # Get tax-exempt transactions
        exemptions = TaxExemption.objects.filter(
            store=report.store,
            exemption_date__gte=report.period_start,
            exemption_date__lte=report.period_end,
        )

        exempt_amount = sum(e.amount for e in exemptions) or Decimal(0)

        # Update report
        report.total_sales = total_sales
        report.total_vat_collected = total_vat
        report.total_refunds = total_refunds
        report.refund_vat = refund_vat
        report.calculate_vat_figures()
        report.save()

    @staticmethod
    def finalize_report(report: VATReport) -> bool:
        """
        Finalize report (lock from further editing).

        Returns:
            True if successful
        """
        if report.status != VATReport.REPORT_STATUS_DRAFT:
            raise ValueError("Only draft reports can be finalized")

        report.status = VATReport.REPORT_STATUS_FINALIZED
        report.save()

        return True

    @staticmethod
    def export_to_csv(report: VATReport) -> str:
        """Export report to CSV format."""
        output = StringIO()
        writer = csv.writer(output)

        # Header
        writer.writerow(["VAT Report", report.report_number])
        writer.writerow(["Period", f"{report.period_start} to {report.period_end}"])
        writer.writerow(["Store", report.store.name])
        writer.writerow([])

        # Summary
        writer.writerow(["Summary"])
        writer.writerow(["Total Sales (including VAT)", report.total_sales])
        writer.writerow(["Total VAT Collected", report.total_vat_collected])
        writer.writerow(["Total VAT Paid (Input)", report.total_vat_paid])
        writer.writerow(["VAT Payable", report.vat_payable])
        writer.writerow(["Total Refunds", report.total_refunds])
        writer.writerow(["Refund VAT", report.refund_vat])
        writer.writerow([])

        # Transactions
        writer.writerow(["Detailed Transactions"])
        writer.writerow(
            [
                "Invoice",
                "Date",
                "Type",
                "Amount (ex-VAT)",
                "VAT",
                "Amount (inc-VAT)",
                "Customer",
                "Payment",
            ]
        )

        for txn in report.transactions.all():
            writer.writerow(
                [
                    txn.invoice_number,
                    txn.transaction_date,
                    txn.get_transaction_type_display(),
                    txn.amount_ex_vat,
                    txn.vat_amount,
                    txn.amount_inc_vat,
                    txn.get_customer_type_display(),
                    txn.payment_method,
                ]
            )

        return output.getvalue()

    @staticmethod
    def export_to_excel(report: VATReport) -> bytes:
        """Export report to Excel format."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "VAT Report"

        # Add title
        ws["A1"] = "VAT Report"
        ws["A1"].font = Font(size=14, bold=True)

        ws["A2"] = report.report_number
        ws["A3"] = f"Period: {report.period_start} to {report.period_end}"
        ws["A4"] = f"Store: {report.store.name}"

        # Summary section
        row = 6
        ws[f"A{row}"] = "Summary"
        ws[f"A{row}"].font = Font(bold=True)

        row += 1
        ws[f"A{row}"] = "Total Sales (including VAT)"
        ws[f"B{row}"] = report.total_sales

        row += 1
        ws[f"A{row}"] = "Total VAT Collected"
        ws[f"B{row}"] = report.total_vat_collected

        row += 1
        ws[f"A{row}"] = "VAT Payable"
        ws[f"B{row}"] = report.vat_payable
        ws[f"B{row}"].font = Font(bold=True)

        # Transactions section
        row += 3
        ws[f"A{row}"] = "Transactions"
        ws[f"A{row}"].font = Font(bold=True)

        row += 1
        headers = [
            "Invoice",
            "Date",
            "Type",
            "Amount ex-VAT",
            "VAT",
            "Amount inc-VAT",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)

        row += 1
        for txn in report.transactions.all():
            ws[f"A{row}"] = txn.invoice_number
            ws[f"B{row}"] = txn.transaction_date
            ws[f"C{row}"] = txn.get_transaction_type_display()
            ws[f"D{row}"] = float(txn.amount_ex_vat)
            ws[f"E{row}"] = float(txn.vat_amount)
            ws[f"F{row}"] = float(txn.amount_inc_vat)
            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 15

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def submit_report(report: VATReport) -> dict:
        """
        Submit report to tax authority (ZATCA).

        Returns:
            {"status": "success|error", ...}
        """
        if not report.can_submit():
            return {
                "status": "error",
                "error": "Report must be finalized before submission",
            }

        try:
            # Generate CSV export
            csv_content = VATReportService.export_to_csv(report)
            csv_file = ContentFile(csv_content.encode())
            report.csv_file.save(
                f"{report.report_number}.csv",
                csv_file,
                save=False,
            )

            # Prepare submission payload
            submission = {
                "report_number": report.report_number,
                "period_start": str(report.period_start),
                "period_end": str(report.period_end),
                "total_sales": str(report.total_sales),
                "total_vat_collected": str(report.total_vat_collected),
                "total_vat_paid": str(report.total_vat_paid),
                "vat_payable": str(report.vat_payable),
                "transactions": [
                    {
                        "invoice": txn.invoice_number,
                        "date": str(txn.transaction_date),
                        "type": txn.transaction_type,
                        "amount_ex_vat": str(txn.amount_ex_vat),
                        "vat": str(txn.vat_amount),
                        "amount_inc_vat": str(txn.amount_inc_vat),
                    }
                    for txn in report.transactions.all()
                ],
            }

            report.submission_payload = submission
            report.submission_number = f"SUB-{timezone.now().timestamp()}"
            report.submitted_at = timezone.now()
            report.status = VATReport.REPORT_STATUS_SUBMITTED
            report.save()

            return {
                "status": "success",
                "submission_number": report.submission_number,
                "message": "Report submitted to tax authority",
            }

        except Exception as e:
            return {
                "status": "error",
                "error": str(e),
            }

    @staticmethod
    def get_monthly_summary(store: Store, year: int, month: int) -> dict:
        """
        Get summary of VAT for month (for quick reference).

        Returns:
            Summary dict
        """
        report = VATReport.objects.filter(
            store=store,
            period_start__year=year,
            period_start__month=month,
        ).first()

        if not report:
            return {}

        return {
            "report_id": report.id,
            "report_number": report.report_number,
            "total_sales": report.total_sales,
            "total_vat_collected": report.total_vat_collected,
            "vat_payable": report.vat_payable,
            "status": report.status,
        }
